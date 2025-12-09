"""
Generador de reportes en formato Excel con formato condicional.
"""
import io
import logging
from datetime import datetime
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Generador de reportes Excel con formato profesional y gráficos.
    """

    def __init__(self, company_name: str = "SOA Minimarket"):
        """
        Inicializa el generador de Excel.

        Args:
            company_name: Nombre de la empresa
        """
        self.company_name = company_name

        # Estilos
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        self.title_font = Font(bold=True, size=16, color="1E40AF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _apply_header_style(self, ws, row: int, max_col: int):
        """
        Aplica estilo al encabezado de una tabla.

        Args:
            ws: Worksheet
            row: Número de fila del encabezado
            max_col: Número de columnas
        """
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _apply_data_style(self, ws, start_row: int, end_row: int, max_col: int):
        """
        Aplica estilo a los datos de una tabla.

        Args:
            ws: Worksheet
            start_row: Fila inicial
            end_row: Fila final
            max_col: Número de columnas
        """
        for row in range(start_row, end_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Fondo alternado
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    def _auto_adjust_columns(self, ws):
        """
        Ajusta automáticamente el ancho de las columnas.

        Args:
            ws: Worksheet
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Max 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width

    def _write_metadata(self, ws, report_title: str, period: str, row: int = 1) -> int:
        """
        Escribe metadata del reporte en el worksheet.

        Args:
            ws: Worksheet
            report_title: Título del reporte
            period: Periodo del reporte
            row: Fila donde escribir

        Returns:
            Siguiente fila disponible
        """
        # Título
        ws.cell(row=row, column=1, value=f"{self.company_name} - {report_title}")
        ws.cell(row=row, column=1).font = self.title_font
        row += 1

        # Fecha generación
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws.cell(row=row, column=1, value=f"Generado: {now}")
        row += 1

        # Periodo
        ws.cell(row=row, column=1, value=f"Periodo: {period.capitalize()}")
        row += 2  # Espacio extra

        return row

    def _write_summary(self, ws, summary_data: Dict[str, Any], row: int) -> int:
        """
        Escribe tabla de resumen.

        Args:
            ws: Worksheet
            summary_data: Datos del resumen
            row: Fila donde escribir

        Returns:
            Siguiente fila disponible
        """
        # Título de sección
        ws.cell(row=row, column=1, value="Resumen")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1

        # Headers
        ws.cell(row=row, column=1, value="Métrica")
        ws.cell(row=row, column=2, value="Valor")
        self._apply_header_style(ws, row, 2)
        row += 1

        start_data_row = row

        # Datos
        for key, value in summary_data.items():
            display_key = key.replace('_', ' ').title()
            ws.cell(row=row, column=1, value=display_key)

            # Formatear valor
            if isinstance(value, float):
                if 'porcentaje' in key.lower() or 'growth' in key.lower():
                    ws.cell(row=row, column=2, value=value).number_format = '0.00"%"'
                else:
                    ws.cell(row=row, column=2, value=value).number_format = '"S/ "#,##0.00'
            elif isinstance(value, int):
                ws.cell(row=row, column=2, value=value).number_format = '#,##0'
            else:
                ws.cell(row=row, column=2, value=str(value))

            row += 1

        # Aplicar estilos
        self._apply_data_style(ws, start_data_row, row - 1, 2)
        row += 1

        return row

    def _write_data_table(self, ws, data: List[Dict[str, Any]], columns: List[str], title: str, row: int) -> int:
        """
        Escribe tabla de datos.

        Args:
            ws: Worksheet
            data: Lista de diccionarios con datos
            columns: Lista de columnas a mostrar
            title: Título de la tabla
            row: Fila donde escribir

        Returns:
            Siguiente fila disponible
        """
        # Título
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1

        if not data:
            ws.cell(row=row, column=1, value="No hay datos disponibles")
            return row + 2

        # Headers
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=row, column=col_idx, value=col.replace('_', ' ').title())

        self._apply_header_style(ws, row, len(columns))
        row += 1

        start_data_row = row

        # Datos
        for item in data:
            for col_idx, col in enumerate(columns, start=1):
                value = item.get(col, '')

                # Formatear según tipo
                if isinstance(value, float):
                    if 'porcentaje' in col.lower():
                        ws.cell(row=row, column=col_idx, value=value).number_format = '0.00"%"'
                    elif any(keyword in col.lower() for keyword in ['precio', 'total', 'revenue', 'gastado', 'ventas', 'ticket']):
                        ws.cell(row=row, column=col_idx, value=value).number_format = '"S/ "#,##0.00'
                    else:
                        ws.cell(row=row, column=col_idx, value=value).number_format = '0.00'
                elif isinstance(value, int):
                    ws.cell(row=row, column=col_idx, value=value).number_format = '#,##0'
                else:
                    ws.cell(row=row, column=col_idx, value=str(value))

            row += 1

        # Aplicar estilos
        self._apply_data_style(ws, start_data_row, row - 1, len(columns))
        row += 1

        return row

    def generate_sales_report(self, report_data: Dict[str, Any], period: str) -> io.BytesIO:
        """
        Genera Excel del reporte de ventas.

        Args:
            report_data: Datos del reporte
            period: Periodo del reporte

        Returns:
            BytesIO con el Excel generado
        """
        buffer = io.BytesIO()
        wb = Workbook()

        try:
            ws = wb.active
            ws.title = "Reporte de Ventas"

            # Metadata
            row = self._write_metadata(ws, "Reporte de Ventas", period)

            # Resumen
            data = report_data.get('data', {})
            summary = {
                'Total Revenue': data.get('total_revenue', 0),
                'Total Pedidos': data.get('total_orders', 0),
                'Ticket Promedio': data.get('average_ticket', 0)
            }
            row = self._write_summary(ws, summary, row)

            # Tabla detallada
            sales_data = data.get('data', [])
            if sales_data:
                columns = ['periodo', 'total_ventas', 'numero_pedidos', 'ticket_promedio']
                row = self._write_data_table(ws, sales_data, columns, "Ventas por Día", row)

            # Ajustar columnas
            self._auto_adjust_columns(ws)

            # Guardar
            wb.save(buffer)
            buffer.seek(0)

            logger.info("✅ Excel de ventas generado exitosamente")
            return buffer

        except Exception as e:
            logger.error(f"Error generando Excel de ventas: {e}")
            raise

    def generate_products_report(self, report_data: Dict[str, Any], period: str) -> io.BytesIO:
        """
        Genera Excel del reporte de productos.

        Args:
            report_data: Datos del reporte
            period: Periodo del reporte

        Returns:
            BytesIO con el Excel generado
        """
        buffer = io.BytesIO()
        wb = Workbook()

        try:
            ws = wb.active
            ws.title = "Top Productos"

            # Metadata
            row = self._write_metadata(ws, "Reporte de Productos Más Vendidos", period)

            # Resumen
            data = report_data.get('data', {})
            summary = {
                'Total Productos': data.get('total_products', 0),
                'Total Unidades': data.get('total_units_sold', 0)
            }
            row = self._write_summary(ws, summary, row)

            # Tabla
            products_data = data.get('data', [])
            if products_data:
                columns = ['nombre', 'categoria', 'total_vendido', 'revenue', 'porcentaje_ventas']
                row = self._write_data_table(ws, products_data, columns, "Top 10 Productos", row)

            # Ajustar columnas
            self._auto_adjust_columns(ws)

            wb.save(buffer)
            buffer.seek(0)

            logger.info("✅ Excel de productos generado exitosamente")
            return buffer

        except Exception as e:
            logger.error(f"Error generando Excel de productos: {e}")
            raise

    def generate_customers_report(self, report_data: Dict[str, Any], period: str) -> io.BytesIO:
        """
        Genera Excel del reporte de clientes.

        Args:
            report_data: Datos del reporte
            period: Periodo del reporte

        Returns:
            BytesIO con el Excel generado
        """
        buffer = io.BytesIO()
        wb = Workbook()

        try:
            ws = wb.active
            ws.title = "Top Clientes"

            # Metadata
            row = self._write_metadata(ws, "Reporte de Clientes Top", period)

            # Resumen
            data = report_data.get('data', {})
            summary = {
                'Total Clientes': data.get('total_customers', 0)
            }
            row = self._write_summary(ws, summary, row)

            # Tabla
            customers_data = data.get('data', [])
            if customers_data:
                columns = ['nombre_completo', 'cantidad_pedidos', 'total_gastado', 'ticket_promedio']
                row = self._write_data_table(ws, customers_data, columns, "Top 20 Clientes", row)

            # Ajustar columnas
            self._auto_adjust_columns(ws)

            wb.save(buffer)
            buffer.seek(0)

            logger.info("✅ Excel de clientes generado exitosamente")
            return buffer

        except Exception as e:
            logger.error(f"Error generando Excel de clientes: {e}")
            raise

    def generate_complete_report(self, reports: Dict[str, Any], period: str) -> io.BytesIO:
        """
        Genera Excel completo con múltiples hojas.

        Args:
            reports: Diccionario con múltiples reportes
            period: Periodo

        Returns:
            BytesIO con el Excel generado
        """
        buffer = io.BytesIO()
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto

        try:
            # Hoja de resumen
            ws_summary = wb.create_sheet("Resumen")
            row = self._write_metadata(ws_summary, "Reporte Completo", period, 1)

            # Ventas
            if 'ventas' in reports:
                ws_ventas = wb.create_sheet("Ventas")
                self.generate_sales_report(reports['ventas'], period)

            # Productos
            if 'productos' in reports:
                ws_productos = wb.create_sheet("Productos")
                self.generate_products_report(reports['productos'], period)

            # Clientes
            if 'clientes' in reports:
                ws_clientes = wb.create_sheet("Clientes")
                self.generate_customers_report(reports['clientes'], period)

            wb.save(buffer)
            buffer.seek(0)

            logger.info("✅ Excel completo generado exitosamente")
            return buffer

        except Exception as e:
            logger.error(f"Error generando Excel completo: {e}")
            raise
